import io
import os
from collections import defaultdict
import openpyxl

from database.connection import get_db_connection

TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates", "financeira")

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}


def obter_sessoes_honorario(user_id: str, projeto_id: int, mes: int, ano: int) -> list:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT a.id,
                       a.data_hora,
                       a.duracao_minutos,
                       a.is_autonomous,
                       a.tipo_atividade,
                       ta.nome AS atividade_nome
                FROM aulas a
                LEFT JOIN mentores m ON m.id = a.mentor_id
                LEFT JOIN turma_atividades ta ON ta.uuid = a.atividade_uuid
                WHERE a.projeto_id = %s
                  AND (
                      (a.is_autonomous = FALSE AND m.user_id = %s)
                      OR (a.is_autonomous = TRUE AND a.responsavel_user_id = %s)
                  )
                  AND (
                      a.estado IN ('confirmada', 'terminada', 'concluida')
                      OR (a.is_autonomous = TRUE AND a.is_realized = TRUE)
                  )
                  AND EXTRACT(month FROM a.data_hora AT TIME ZONE 'Europe/Lisbon') = %s
                  AND EXTRACT(year FROM a.data_hora AT TIME ZONE 'Europe/Lisbon') = %s
                ORDER BY a.data_hora
            """, (projeto_id, user_id, user_id, mes, ano))
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, row)) for row in cur.fetchall()]
    finally:
        conn.close()


def agrupar_por_atividade(sessoes: list, valor_hora: float) -> list:
    grupos: dict = defaultdict(float)
    for s in sessoes:
        chave = s.get("atividade_nome") or s.get("tipo_atividade") or "Outras atividades"
        grupos[chave] += (s.get("duracao_minutos") or 0) / 60.0

    if len(grupos) > 15:
        raise ValueError(
            f"Demasiadas categorias de atividade ({len(grupos)}) para o mês selecionado. "
            "O máximo são 15 linhas por nota de honorário."
        )

    return [
        {
            "descricao": desc,
            "horas": round(horas, 2),
            "valor_hora": valor_hora,
            "valor": round(horas * valor_hora, 2),
        }
        for desc, horas in grupos.items()
    ]


def obter_dados_prestador(user_id: str, projeto_id: int) -> dict:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT p.id,
                       p.full_name,
                       p.nif,
                       p.morada,
                       p.cod_postal,
                       p.funcao,
                       upr.valor_hora
                FROM profiles p
                LEFT JOIN user_projeto_rates upr
                       ON upr.user_id = p.id AND upr.projeto_id = %s
                WHERE p.id = %s
            """, (projeto_id, user_id))
            row = cur.fetchone()
            if not row:
                raise ValueError("Utilizador não encontrado.")
            cols = [d[0] for d in cur.description]
            return dict(zip(cols, row))
    finally:
        conn.close()


def obter_projeto_config_honorario(projeto_id: int) -> dict:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, nome, codigo_projeto,
                       usa_template_pis,
                       honorario_entidade, honorario_morada, honorario_cod_postal,
                       honorario_nipc, honorario_designacao
                FROM projetos
                WHERE id = %s
            """, (projeto_id,))
            row = cur.fetchone()
            if not row:
                raise ValueError("Projeto não encontrado.")
            cols = [d[0] for d in cur.description]
            return dict(zip(cols, row))
    finally:
        conn.close()


def _preencher_pis(wb, prestador: dict, _projeto: dict, grupos: list, mes: int, data_emissao: str):
    ws = wb["Nota Honorário PIS"]

    label_mes = f"{MESES_PT.get(mes, str(mes))} {data_emissao[:4]}"
    ws["I12"] = label_mes
    ws["K12"] = data_emissao

    ws["C13"] = prestador.get("full_name") or ""
    ws["C14"] = prestador.get("morada") or ""
    ws["D15"] = prestador.get("cod_postal") or ""
    ws["D16"] = prestador.get("nif") or ""
    ws["C18"] = prestador.get("funcao") or ""

    for i, grupo in enumerate(grupos[:15]):
        row = 21 + i
        ws[f"C{row}"] = grupo["descricao"]
        ws[f"I{row}"] = grupo["horas"]
        ws[f"J{row}"] = grupo["valor_hora"]
        # K column already has =IFERROR(IF(I{row}="",ROUND(1*J{row},2),ROUND(I{row}*J{row},2)),"")


def _preencher_gulbenkian(wb, prestador: dict, projeto: dict, grupos: list, mes: int, data_emissao: str):
    ws = wb["Folha1"]

    # Entity header (configurable per project)
    ws["D15"] = projeto.get("honorario_entidade") or ""
    ws["D16"] = projeto.get("honorario_morada") or ""
    ws["D17"] = projeto.get("honorario_cod_postal") or ""
    ws["D18"] = projeto.get("honorario_nipc") or ""

    # Project info
    ws["C21"] = projeto.get("honorario_designacao") or projeto.get("nome") or ""
    ws["K21"] = projeto.get("codigo_projeto") or ""

    # Prestador info
    label_mes = f"{MESES_PT.get(mes, str(mes))} {data_emissao[:4]}"
    ws["I23"] = label_mes
    ws["J23"] = data_emissao
    ws["D24"] = prestador.get("full_name") or ""
    ws["D25"] = prestador.get("morada") or ""
    ws["D26"] = prestador.get("cod_postal") or ""
    ws["D27"] = prestador.get("nif") or ""
    ws["D29"] = prestador.get("funcao") or ""

    # Data rows (no formulas — compute in Python)
    subtotal = 0.0
    for i, grupo in enumerate(grupos[:15]):
        row = 32 + i
        ws[f"C{row}"] = grupo["descricao"]
        ws[f"H{row}"] = grupo["horas"]
        ws[f"I{row}"] = grupo["valor_hora"]
        valor = round(grupo["horas"] * grupo["valor_hora"], 2)
        ws[f"J{row}"] = valor
        subtotal += valor

    subtotal = round(subtotal, 2)
    ws["J47"] = subtotal
    ws["J48"] = 0       # IVA
    ws["J49"] = 0       # IRS
    ws["J50"] = subtotal


def gerar_honorario(
    _user_id: str,
    target_user_id: str,
    projeto_id: int,
    mes: int,
    ano: int,
    data_emissao: str,
) -> bytes:
    projeto = obter_projeto_config_honorario(projeto_id)
    prestador = obter_dados_prestador(target_user_id, projeto_id)

    valor_hora = float(prestador.get("valor_hora") or 0)
    sessoes = obter_sessoes_honorario(target_user_id, projeto_id, mes, ano)

    if not sessoes:
        raise ValueError(
            f"Nenhuma sessão encontrada para {MESES_PT.get(mes, str(mes))} {ano} "
            "neste projeto para o utilizador selecionado."
        )

    grupos = agrupar_por_atividade(sessoes, valor_hora)

    usa_pis = projeto.get("usa_template_pis", False)
    template_name = "honorario_pis.xlsx" if usa_pis else "honorario_gulbenkian.xlsx"
    template_path = os.path.join(TEMPLATES_DIR, template_name)

    wb = openpyxl.load_workbook(template_path)

    if usa_pis:
        _preencher_pis(wb, prestador, projeto, grupos, mes, data_emissao)
    else:
        _preencher_gulbenkian(wb, prestador, projeto, grupos, mes, data_emissao)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def obter_preview_honorario(target_user_id: str, projeto_id: int, mes: int, ano: int) -> dict:
    prestador = obter_dados_prestador(target_user_id, projeto_id)
    valor_hora = float(prestador.get("valor_hora") or 0)
    sessoes = obter_sessoes_honorario(target_user_id, projeto_id, mes, ano)
    grupos = agrupar_por_atividade(sessoes, valor_hora) if sessoes else []
    subtotal = round(sum(g["valor"] for g in grupos), 2)
    return {
        "grupos": grupos,
        "subtotal": subtotal,
        "total_horas": round(sum(g["horas"] for g in grupos), 2),
        "valor_hora": valor_hora,
        "prestador_nome": prestador.get("full_name") or "",
        "mes": mes,
        "ano": ano,
        "num_sessoes": len(sessoes),
    }


# ─── Rates management ────────────────────────────────────────────────────────

def listar_rates_projeto(projeto_id: int) -> list:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT p.id, p.full_name, p.role,
                       COALESCE(upr.valor_hora, 0) AS valor_hora
                FROM profiles p
                LEFT JOIN user_projeto_rates upr ON upr.user_id = p.id AND upr.projeto_id = %s
                LEFT JOIN user_project_access upa ON upa.user_id = p.id AND upa.projeto_id = %s
                WHERE p.is_root = TRUE
                   OR p.project_scoped = FALSE
                   OR upa.user_id IS NOT NULL
                ORDER BY p.full_name
            """, (projeto_id, projeto_id))
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, row)) for row in cur.fetchall()]
    finally:
        conn.close()


def obter_rate(user_id: str, projeto_id: int) -> dict:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COALESCE(valor_hora, 0) AS valor_hora
                FROM user_projeto_rates
                WHERE user_id = %s AND projeto_id = %s
            """, (user_id, projeto_id))
            row = cur.fetchone()
            return {"valor_hora": float(row[0]) if row else 0.0}
    finally:
        conn.close()


def upsert_rate(user_id: str, projeto_id: int, valor_hora: float) -> dict:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO user_projeto_rates (user_id, projeto_id, valor_hora)
                VALUES (%s, %s, %s)
                ON CONFLICT (user_id, projeto_id) DO UPDATE SET valor_hora = EXCLUDED.valor_hora
                RETURNING id, user_id, projeto_id, valor_hora
            """, (user_id, projeto_id, valor_hora))
            row = cur.fetchone()
            cols = [d[0] for d in cur.description]
            conn.commit()
            return dict(zip(cols, row))
    finally:
        conn.close()
