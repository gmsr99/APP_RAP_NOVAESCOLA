import io
import os
import datetime
import requests
from calendar import monthrange
from collections import defaultdict, OrderedDict

import openpyxl

from database.connection import get_db_connection

KM_TEMPLATES_DIR = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "frontend", "public", "templates", "kms"
)

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

TAXA_KM = 0.30


def obter_projeto_config_km(projeto_id: int) -> dict:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, nome, codigo_projeto, usar_template_km_proprio
                FROM projetos WHERE id = %s
            """, (projeto_id,))
            row = cur.fetchone()
            if not row:
                raise ValueError("Projeto não encontrado.")
            cols = [d[0] for d in cur.description]
            return dict(zip(cols, row))
    finally:
        conn.close()


def obter_dados_mentor_km(user_id: str) -> dict:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT p.full_name, p.nif, p.morada, p.matricula_viatura,
                       m.latitude, m.longitude
                FROM profiles p
                LEFT JOIN mentores m ON m.user_id = p.id
                WHERE p.id = %s
            """, (user_id,))
            row = cur.fetchone()
            if not row:
                raise ValueError("Utilizador não encontrado.")
            cols = [d[0] for d in cur.description]
            return dict(zip(cols, row))
    finally:
        conn.close()


def obter_sessoes_km(
    user_id: str, projeto_id: int,
    data_inicio: datetime.date, data_fim: datetime.date
) -> list:
    """Retorna sessões presenciais com leva_carro=TRUE no período, ordenadas por dia e hora."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    DATE(a.data_hora AT TIME ZONE 'Europe/Lisbon') AS dia,
                    a.data_hora,
                    e.latitude,
                    e.longitude,
                    e.nome AS estabelecimento,
                    p.nome AS projeto_nome
                FROM aulas a
                JOIN mentores m ON m.id = a.mentor_id
                JOIN turmas t ON t.id = a.turma_id
                JOIN estabelecimentos e ON e.id = t.estabelecimento_id
                JOIN projetos p ON p.id = a.projeto_id
                WHERE a.projeto_id = %s
                  AND a.is_autonomous = FALSE
                  AND m.user_id = %s
                  AND a.leva_carro = TRUE
                  AND a.estado = 'terminada'
                  AND DATE(a.data_hora AT TIME ZONE 'Europe/Lisbon') BETWEEN %s AND %s
                ORDER BY dia, a.data_hora
            """, (projeto_id, user_id, data_inicio, data_fim))
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, row)) for row in cur.fetchall()]
    finally:
        conn.close()


def obter_leva_carro_dia(user_id: str, data: str) -> bool | None:
    """Verifica se o mentor já respondeu 'leva_carro' neste dia (lê de aulas.leva_carro)."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT a.leva_carro
                FROM aulas a
                JOIN mentores m ON m.id = a.mentor_id
                WHERE m.user_id = %s
                  AND a.leva_carro IS NOT NULL
                  AND a.is_autonomous = FALSE
                  AND DATE(a.data_hora AT TIME ZONE 'Europe/Lisbon') = %s
                LIMIT 1
            """, (user_id, data))
            row = cur.fetchone()
            return row[0] if row else None
    finally:
        conn.close()


def _calcular_rota_diaria(
    mentor_lat: float, mentor_lng: float, waypoints: list[tuple]
) -> float:
    """
    Calcula distância total da rota: Casa → WP1 → WP2 → ... → Casa via OSRM.
    Deduplica waypoints consecutivos idênticos (mentor fica no mesmo local entre sessões).
    Retorna km arredondado a 1 decimal, ou 0.0 se falhar.
    """
    if not waypoints:
        return 0.0

    # Remove consecutive duplicates (same establishment back-to-back)
    unique_wps: list[tuple] = []
    for wp in waypoints:
        if not unique_wps or wp != unique_wps[-1]:
            unique_wps.append(wp)

    all_points = [(mentor_lat, mentor_lng)] + unique_wps + [(mentor_lat, mentor_lng)]
    coords = ";".join(f"{lng},{lat}" for lat, lng in all_points)
    url = f"https://router.project-osrm.org/route/v1/driving/{coords}?overview=false"

    try:
        resp = requests.get(url, timeout=10)
        data = resp.json()
        if data.get("routes"):
            return round(data["routes"][0]["distance"] / 1000, 1)
    except Exception:
        pass
    return 0.0


def _calcular_ciclo(mes: int, ano: int, usa_pis: bool):
    if usa_pis:
        prev_mes = mes - 1 if mes > 1 else 12
        prev_ano = ano if mes > 1 else ano - 1
        data_inicio = datetime.date(prev_ano, prev_mes, 20)
        data_fim = datetime.date(ano, mes, 19)
        data_doc = datetime.date(ano, mes, 20)
    else:
        ultimo_dia = monthrange(ano, mes)[1]
        data_inicio = datetime.date(ano, mes, 1)
        data_fim = datetime.date(ano, mes, ultimo_dia)
        data_doc = data_fim
    return data_inicio, data_fim, data_doc


def _preencher_generico(
    wb, mentor: dict, projeto: dict,
    dias_km: dict, mes: int, ano: int, data_doc: datetime.date
):
    ws = wb.worksheets[0]
    ws["L16"] = MESES_PT.get(mes, str(mes))
    ws["M16"] = f"Ano: {ano}"

    projeto_nome = projeto.get("nome") or ""
    for dia_date, info in dias_km.items():
        row = 21 + dia_date.day
        ws[f"M{row}"] = info["kms"]
        ws[f"I{row}"] = info["motivo"]
        ws[f"K{row}"] = projeto_nome

    ws["D57"] = mentor.get("full_name") or ""
    ws["L57"] = data_doc.strftime("%d/%m/%Y")
    ws["D59"] = mentor.get("matricula_viatura") or ""
    ws["D61"] = mentor.get("nif") or ""


def _preencher_pis(
    wb, mentor: dict, projeto: dict,
    dias_km: dict, mes: int, ano: int, data_doc: datetime.date
):
    ws = wb["MODELO"]
    ws["K7"] = MESES_PT.get(mes, str(mes))
    ws["M7"] = ano

    projeto_nome = projeto.get("nome") or ""
    for dia_date, info in dias_km.items():
        dia = dia_date.day
        if dia_date.month != mes:
            # Previous month: days 20-31 → rows 12-23
            row = 12 + (dia - 20)
        else:
            # Current month: days 1-19 → rows 24-42
            row = 24 + (dia - 1)
        if 12 <= row <= 42:
            ws[f"M{row}"] = info["kms"]
            ws[f"I{row}"] = info["motivo"]
            ws[f"K{row}"] = projeto_nome

    ws["D47"] = mentor.get("full_name") or ""
    ws["L47"] = data_doc.strftime("%d/%m/%Y")
    ws["D48"] = mentor.get("matricula_viatura") or ""
    ws["D49"] = mentor.get("nif") or ""


def gerar_mapa_kms(user_id: str, projeto_id: int, mes: int, ano: int) -> bytes:
    projeto = obter_projeto_config_km(projeto_id)
    mentor = obter_dados_mentor_km(user_id)

    usa_pis = projeto.get("usar_template_km_proprio", False)
    data_inicio, data_fim, data_doc = _calcular_ciclo(mes, ano, usa_pis)

    sessoes = obter_sessoes_km(user_id, projeto_id, data_inicio, data_fim)
    if not sessoes:
        raise ValueError(
            f"Nenhuma sessão com carro registada para o período "
            f"{data_inicio.strftime('%d/%m/%Y')} – {data_fim.strftime('%d/%m/%Y')}."
        )

    mentor_lat = mentor.get("latitude")
    mentor_lng = mentor.get("longitude")
    if not mentor_lat or not mentor_lng:
        raise ValueError(
            "O teu perfil não tem localização definida. "
            "Define a tua morada nas definições para calcular os KMs automaticamente."
        )

    # Group by day, build waypoints list
    grupos: dict = defaultdict(lambda: {"waypoints": [], "estabelecimentos": []})
    for s in sessoes:
        dia = s["dia"]
        if s.get("latitude") and s.get("longitude"):
            grupos[dia]["waypoints"].append((float(s["latitude"]), float(s["longitude"])))
        grupos[dia]["estabelecimentos"].append(s["estabelecimento"] or "")

    # Calculate daily distances via OSRM
    dias_km: dict = {}
    for dia, info in sorted(grupos.items()):
        kms = _calcular_rota_diaria(mentor_lat, mentor_lng, info["waypoints"])
        # Unique establishment names preserving first-seen order
        seen: dict = OrderedDict.fromkeys(info["estabelecimentos"])
        motivo = " · ".join(seen)
        dias_km[dia] = {"kms": kms, "motivo": motivo}

    template_name = "RAP NE IMP27_v1 ATB - Mapa kms.xlsx" if usa_pis else "Mapa KM.xlsx"
    template_path = os.path.join(KM_TEMPLATES_DIR, template_name)
    wb = openpyxl.load_workbook(template_path)

    if usa_pis:
        _preencher_pis(wb, mentor, projeto, dias_km, mes, ano, data_doc)
    else:
        _preencher_generico(wb, mentor, projeto, dias_km, mes, ano, data_doc)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def obter_leva_carro_resumo(data_inicio: str, data_fim: str) -> dict:
    """
    Retorna, para cada dia no intervalo, a lista de mentores que confirmaram levar carro.
    Lê de aulas.leva_carro (definido no momento de confirmação da sessão).
    Formato: { "YYYY-MM-DD": [{ "mentor_nome": str, "estabelecimentos": [str] }] }
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    DATE(a.data_hora AT TIME ZONE 'Europe/Lisbon') AS dia,
                    p.full_name AS mentor_nome,
                    ARRAY_AGG(DISTINCT e.nome ORDER BY e.nome) AS estabelecimentos
                FROM aulas a
                JOIN mentores m ON m.id = a.mentor_id
                JOIN profiles p ON p.id = m.user_id
                JOIN turmas t ON t.id = a.turma_id
                JOIN estabelecimentos e ON e.id = t.estabelecimento_id
                WHERE a.leva_carro = TRUE
                  AND a.is_autonomous = FALSE
                  AND a.estado IN ('confirmada', 'terminada')
                  AND DATE(a.data_hora AT TIME ZONE 'Europe/Lisbon') BETWEEN %s AND %s
                GROUP BY dia, p.full_name
                ORDER BY dia, p.full_name
            """, (data_inicio, data_fim))
            result: dict = {}
            for row in cur.fetchall():
                dia_str = row[0].isoformat() if hasattr(row[0], 'isoformat') else str(row[0])
                if dia_str not in result:
                    result[dia_str] = []
                result[dia_str].append({
                    "mentor_nome": row[1],
                    "estabelecimentos": list(row[2]) if row[2] else [],
                })
            return result
    finally:
        conn.close()


def obter_preview_mapa_kms(user_id: str, projeto_id: int, mes: int, ano: int) -> dict:
    """Preview sem gerar o ficheiro: retorna totais e avisos de dados em falta."""
    projeto = obter_projeto_config_km(projeto_id)
    usa_pis = projeto.get("usar_template_km_proprio", False)
    data_inicio, data_fim, data_doc = _calcular_ciclo(mes, ano, usa_pis)
    sessoes = obter_sessoes_km(user_id, projeto_id, data_inicio, data_fim)
    mentor = obter_dados_mentor_km(user_id)

    avisos: list[str] = []
    if not mentor.get("latitude") or not mentor.get("longitude"):
        avisos.append("O teu perfil não tem localização definida.")
    if not mentor.get("matricula_viatura"):
        avisos.append("Matrícula da viatura não definida no perfil.")

    estabs_sem_coords = {s["estabelecimento"] for s in sessoes
                         if not s.get("latitude") or not s.get("longitude")}
    if estabs_sem_coords:
        avisos.append(f"Sem coordenadas: {', '.join(sorted(estabs_sem_coords))}")

    dias: dict = defaultdict(list)
    for s in sessoes:
        dias[s["dia"]].append(s)

    return {
        "num_sessoes": len(sessoes),
        "num_dias": len(dias),
        "data_inicio": data_inicio.isoformat(),
        "data_fim": data_fim.isoformat(),
        "data_doc": data_doc.isoformat(),
        "mentor_nome": mentor.get("full_name") or "",
        "avisos": avisos,
    }
